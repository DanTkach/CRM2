from io import BytesIO
from docxtpl import DocxTemplate
import openpyxl
from django.shortcuts import render, redirect
from django.forms import modelform_factory
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .process import render_to_pdf
from django.views.generic import View
from django.http import HttpResponse, StreamingHttpResponse, JsonResponse
import os
import urllib.request
import xml.etree.ElementTree as ET

# Create your views here.
from .models import *
from .forms import (CreateUserForm, PhysicalClientForm, JuridicalClientForm,
                    ContractForm, PaymentForm, GuarantorForm, PrognosisForm,
                    ArrearsForm, CommentsForm, ReportForm, GajistForm,
                    PaymentsForm)
from .filters import ClientFilter, ArrearsFilter
from .decorators import unauthenticated_user, allowed_users, admin_only
from .functions import create_spread_sheet, verbose_loan, verbose_interest
from datetime import date, datetime
from dateutil.relativedelta import relativedelta


production = False


@login_required(login_url='login')
def credit_calculator(request):
    return render(request, 'accounts/credit_calculator.html')


class ExportXlsx(View):
    def get(self, request, *args, **kwargs):
        directory = os.getcwd()
        global production
        if production:
            path = directory + '/CRM/accounts/templates/pdf/statistics.xlsx'
        else:
            path = directory + '/accounts/templates/pdf/statistics.xlsx'
        wb_obj = openpyxl.load_workbook(path)

        contracts = Contract.objects.all()
        clients = Client.objects.all()
        all_clients_count = clients.count()
        female_clients_count = Client.objects.filter(gender="Female").count()
        physical_clients = []
        juridical_clients = []
        [agri_ph, indu_ph, imob_ph, come_ph, serv_ph,
         cons_ph, alte_ph, agri_ju, indu_ju, imob_ju,
         come_ju, serv_ju, cons_ju, alte_ju] = [0, 0, 0, 0, 0,
                                                0, 0, 0, 0, 0,
                                                0, 0, 0, 0]
        standard = []
        supravegheate = []
        substandard = []
        dubioase = []
        compromise = []
        standard_contracts = [0, 0]
        supravegheate_contracts = [0, 0]
        substandard_contracts = [0, 0]
        dubioase_contracts = [0, 0]
        compromise_contracts = [0, 0]
        busy = []
        for contract in contracts:
            if contract.client.person == "Physical":
                physical_clients.append(contract)
                if contract.sector == "Agricultură":
                    agri_ph += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Industria Alim. și Procesare":
                    indu_ph += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Imobil":
                    imob_ph += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Comerț":
                    come_ph += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Prestarea serviciilor":
                    serv_ph += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Consum":
                    cons_ph += int(contract.loan * contract.exchange_rate)
                else:
                    alte_ph += int(contract.loan * contract.exchange_rate)
            else:
                juridical_clients.append(contract)
                if contract.sector == "Agricultură":
                    agri_ju += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Industria Alim. și Procesare":
                    indu_ju += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Imobil":
                    imob_ju += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Comerț":
                    come_ju += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Prestarea serviciilor":
                    serv_ju += int(contract.loan * contract.exchange_rate)
                elif contract.sector == "Consum":
                    cons_ju += int(contract.loan * contract.exchange_rate)
                else:
                    alte_ju += int(contract.loan * contract.exchange_rate)
            payments = Payment.objects.filter(contract_id=contract.id)
            table, profile, penalties = create_spread_sheet(
                str(contract.date_created).split(" ")[0],
                contract.months,
                contract.annual_payments,
                float(contract.loan),
                float(contract.interest_rate),
                contract.calc_method,
                str(date.today()),
                [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
                float(contract.month_sum),
                contract.grace_period,
                float(contract.fee_issue or 0),
                int(contract.fee_issue_months or 1),
                float(contract.comission or 0)
            )
            if profile[0][21] > 360 or contract.client_id in compromise:
                compromise.append(contract.client_id)
                compromise_contracts[0] += profile[0][10]
                compromise_contracts[1] += profile[0][13]
                busy.append(contract.client_id)
            elif profile[0][21] > 180 or contract.client_id in dubioase:
                dubioase_contracts[0] += profile[0][10]
                dubioase_contracts[1] += profile[0][13]
                dubioase.append(contract.client_id)
                busy.append(contract.client_id)
            elif profile[0][21] > 90 or contract.client_id in substandard:
                substandard_contracts[0] += profile[0][10]
                substandard_contracts[1] += profile[0][13]
                substandard.append(contract.client_id)
                busy.append(contract.client_id)
            elif profile[0][21] > 30 or contract.client_id in supravegheate:
                supravegheate_contracts[0] += profile[0][10]
                supravegheate_contracts[1] += profile[0][13]
                supravegheate.append(contract.client_id)
                busy.append(contract.client_id)
            elif profile[0][21] < 31 or contract.client_id in standard:
                standard_contracts[0] += profile[0][10]
                standard_contracts[1] += profile[0][13]
                standard.append(contract.client_id)

        sheet_10 = wb_obj.get_sheet_by_name('10_DG')
        c1 = sheet_10.cell(row=16, column=4)
        c1.value = str(all_clients_count)
        c2 = sheet_10.cell(row=17, column=4)
        c2.value = str(female_clients_count)
        sheet_11 = wb_obj.get_sheet_by_name('11_DU')
        c3 = sheet_11.cell(row=9, column=7)
        c3.value = '{:,}'.format(agri_ju).replace(',', '.').replace('.00', '')
        c4 = sheet_11.cell(row=10, column=7)
        c4.value = '{:,}'.format(indu_ju).replace(',', '.').replace('.00', '')
        c5 = sheet_11.cell(row=11, column=7)
        c5.value = '{:,}'.format(imob_ju).replace(',', '.').replace('.00', '')
        c6 = sheet_11.cell(row=12, column=7)
        c6.value = '{:,}'.format(come_ju).replace(',', '.').replace('.00', '')
        c7 = sheet_11.cell(row=13, column=7)
        c7.value = '{:,}'.format(serv_ju).replace(',', '.').replace('.00', '')
        c8 = sheet_11.cell(row=14, column=7)
        c8.value = '{:,}'.format(cons_ju).replace(',', '.').replace('.00', '')
        c9 = sheet_11.cell(row=15, column=7)
        c9.value = '{:,}'.format(alte_ju).replace(',', '.').replace('.00', '')
        c10 = sheet_11.cell(row=17, column=7)
        c10.value = '{:,}'.format(agri_ph).replace(',', '.').replace('.00', '')
        c11 = sheet_11.cell(row=18, column=7)
        c11.value = '{:,}'.format(indu_ph).replace(',', '.').replace('.00', '')
        c12 = sheet_11.cell(row=19, column=7)
        c12.value = '{:,}'.format(imob_ph).replace(',', '.').replace('.00', '')
        c13 = sheet_11.cell(row=20, column=7)
        c13.value = '{:,}'.format(come_ph).replace(',', '.').replace('.00', '')
        c14 = sheet_11.cell(row=21, column=7)
        c14.value = '{:,}'.format(serv_ph).replace(',', '.').replace('.00', '')
        c15 = sheet_11.cell(row=22, column=7)
        c15.value = '{:,}'.format(cons_ph).replace(',', '.').replace('.00', '')
        c16 = sheet_11.cell(row=23, column=7)
        c16.value = '{:,}'.format(alte_ph).replace(',', '.').replace('.00', '')

        sheet_8 = wb_obj.get_sheet_by_name('8_PN')
        b11 = sheet_8.cell(row=9, column=5)
        b11.value = str(int(standard_contracts[0]))
        b12 = sheet_8.cell(row=9, column=6)
        b12.value = str(int(standard_contracts[1]))
        b21 = sheet_8.cell(row=10, column=5)
        b21.value = str(int(supravegheate_contracts[0]))
        b22 = sheet_8.cell(row=10, column=6)
        b22.value = str(int(supravegheate_contracts[1]))
        b31 = sheet_8.cell(row=11, column=5)
        b31.value = str(int(substandard_contracts[0]))
        b32 = sheet_8.cell(row=11, column=6)
        b32.value = str(int(substandard_contracts[1]))
        b41 = sheet_8.cell(row=12, column=5)
        b41.value = str(int(dubioase_contracts[0]))
        b42 = sheet_8.cell(row=12, column=6)
        b42.value = str(int(dubioase_contracts[1]))
        b51 = sheet_8.cell(row=13, column=5)
        b51.value = str(int(compromise_contracts[0]))
        b52 = sheet_8.cell(row=13, column=6)
        b52.value = str(int(compromise_contracts[1]))

        buffer = BytesIO()
        wb_obj.save(buffer)  # save your memory stream
        buffer.seek(0)

        response = StreamingHttpResponse(
            streaming_content=buffer,  # use the stream's content
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

        response['Content-Disposition'] = 'attachment;filename=Statistics.xlsx'
        response["Content-Encoding"] = 'UTF-8'

        return response


class ExportDocx(View):
    def get(self, request, pk, *args, **kwargs):
        # create an empty document object
        directory = os.getcwd()
        global production
        if production:
            document_dir = directory + '/CRM/accounts/templates/pdf/'
        else:
            document_dir = directory + '/accounts/templates/pdf/'
        contract = Contract.objects.get(id=pk)
        client = Client.objects.get(id=contract.client_id)
        if client.person == "Physical":
            document = DocxTemplate(document_dir + 'template.docx')
        else:
            document = DocxTemplate(
                document_dir + 'template_jur_contract.docx')
        payments = Payment.objects.filter(contract_id=pk)
        guarantors = contract.fidejusor_set.all()
        lienholders = contract.gajist_set.all()
        table, profile, penalties = create_spread_sheet(
            str(contract.date_created).split(" ")[0],
            contract.months,
            contract.annual_payments,
            float(contract.loan),
            float(contract.interest_rate),
            contract.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract.month_sum),
            contract.grace_period,
            float(contract.fee_issue or 0),
            int(contract.fee_issue_months or 1),
            float(contract.comission or 0)
        )
        ptotal = 0
        itotal = 0
        stotal = 0
        tbl_contents = []
        for row in table:
            tbl_contents.append(
                {'nr': row[0],
                 'date': str(row[1]).replace('/', '.'),
                 'days': row[13],
                 'sum': "{:.2f}".format(row[2]),
                 'rate': "{:.2f}".format(row[3]),
                 'interest': "{:.2f}".format(row[12]),
                 'total': "{:.2f}".format(row[5])}
            )
            ptotal += row[3]
            itotal += row[12]
            stotal += row[5]

        char = 'b'
        guarantor_txt = ''
        for guarantor in guarantors:
            guarantor_txt += char + ")    Contract de fidejusiune nr. " + \
                str(contract.id) + "-"
            guarantor_txt += str(guarantor.num) + " cu " + \
                str(guarantor.last_name) + " "
            guarantor_txt += str(guarantor.first_name) + \
                " (IDNP " + str(guarantor.idnp) + ")\n"
            char = chr(ord(char) + 1)
        lienholder_txt = ''
        for lienholder in lienholders:
            lienholder_txt += char + ")    Contract de gaj nr. " + \
                str(contract.id) + "-"
            lienholder_txt += str(lienholder.num) + " cu " + \
                str(lienholder.last_name) + " "
            lienholder_txt += str(lienholder.first_name) + \
                " (IDNP " + str(lienholder.idnp) + ")\n"
            char = chr(ord(char) + 1)
        if client.idnp:
            if client.person == "Physical":
                context = {
                    'id': pk,
                    'contract_date': str(date.fromisoformat(str(contract.date_created)).strftime("%d.%m.%Y")),
                    'first_name': client.first_name,
                    'last_name': client.last_name,
                    'company_name': client.company_name,
                    'jur_adress': client.jur_address,
                    'admin_name': client.administrator,
                    'reg_nr': client.registration_nr,
                    'location': client.location,
                    'idnp': client.idnp,
                    'birth_date': str(date.fromisoformat(str(client.birth_date)).strftime("%d.%m.%Y")),
                    'loan_sum': "{:.2f}".format(contract.loan),
                    'loan_sum_verbose': verbose_loan(round(float(contract.loan), 2)),
                    'scope': contract.scope,
                    'last_payment_date': str(table[len(table)-1][1]).replace('/', '.'),
                    'interest_rate': "{:.2f}".format(contract.interest_rate),
                    'interest_rate_verbose': verbose_interest(contract.interest_rate),
                    'phone_num': client.phone_num,
                    'bank_name': client.bank_name,
                    'bank_code': client.bank_code,
                    'bank_account': client.bank_account,
                    'months': contract.months,
                    'payments': str(int(contract.months / int(12 / contract.annual_payments))),
                    'tbl_contents': tbl_contents,
                    'guarantor_txt': guarantor_txt,
                    'lienholder_txt': lienholder_txt,
                    'ptotal': "{:.2f}".format(ptotal),
                    'itotal': "{:.2f}".format(itotal),
                    'stotal': "{:.2f}".format(stotal),
                    'perc': "{:.2f}".format(contract.comission).replace('.', ','),
                    'commission': "{:.2f}".format(contract.comission / 100 * contract.loan)
                }
            else:
                context = {
                    'id': pk,
                    'contract_date': str(date.fromisoformat(str(contract.date_created)).strftime("%d.%m.%Y")),
                    'first_name': client.first_name,
                    'last_name': client.last_name,
                    'company_name': client.company_name,
                    'jur_adress': client.jur_address,
                    'admin_name': client.administrator,
                    'reg_date': str(date.fromisoformat(str(client.registration_date)).strftime("%d.%m.%Y")),
                    'reg_nr': client.registration_nr,
                    'location': client.location,
                    'idnp': client.idnp,
                    'loan_sum': "{:.2f}".format(contract.loan),
                    'loan_sum_verbose': verbose_loan(round(float(contract.loan), 2)),
                    'scope': contract.scope,
                    'last_payment_date': str(table[len(table)-1][1]).replace('/', '.'),
                    'interest_rate': "{:.2f}".format(contract.interest_rate),
                    'interest_rate_verbose': verbose_interest(contract.interest_rate),
                    'phone_num': client.phone_num,
                    'bank_name': client.bank_name,
                    'bank_code': client.bank_code,
                    'bank_account': client.bank_account,
                    'months': contract.months,
                    'payments': str(int(contract.months / int(12 / contract.annual_payments))),
                    'tbl_contents': tbl_contents,
                    'guarantor_txt': guarantor_txt,
                    'lienholder_txt': lienholder_txt,
                    'ptotal': "{:.2f}".format(ptotal),
                    'itotal': "{:.2f}".format(itotal),
                    'stotal': "{:.2f}".format(stotal),
                    'perc': "{:.2f}".format(contract.comission).replace('.', ','),
                    'commission': "{:.2f}".format(contract.comission / 100 * contract.loan)
                }
            document.render(context)
            buffer = BytesIO()
            document.save(buffer)
            buffer.seek(0)
            response = StreamingHttpResponse(
                streaming_content=buffer,  # use the stream's content
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
            response['Content-Disposition'] = 'attachment;filename=Contract_' + \
                str(contract.id) + '.docx'
            response["Content-Encoding"] = 'UTF-8'
            return response
        else:
            messages.error(request, "Client lacks info")
            return redirect('/view_contract/' + str(pk) + '/')


class ExportDocxChestionar(View):
    def get(self, request, pk, *args, **kwargs):
        # create an empty document object
        directory = os.getcwd()
        global production
        if production:
            document = DocxTemplate(
                directory + '/CRM/accounts/templates/pdf/chestionar.docx')
        else:
            document = DocxTemplate(
                directory + '/accounts/templates/pdf/chestionar.docx')
        contract = Contract.objects.get(id=pk)
        client = Client.objects.get(id=contract.client_id)
        payments = Payment.objects.filter(contract_id=pk)
        guarantors = contract.fidejusor_set.all()
        table, profile, penalties = create_spread_sheet(
            str(contract.date_created).split(" ")[0],
            contract.months,
            contract.annual_payments,
            float(contract.loan),
            float(contract.interest_rate),
            contract.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract.month_sum),
            contract.grace_period,
            float(contract.fee_issue or 0),
            int(contract.fee_issue_months or 1),
            float(contract.comission or 0)
        )
        ptotal = 0
        itotal = 0
        stotal = 0
        tbl_contents = []
        for row in table:
            tbl_contents.append(
                {'nr': row[0],
                 'date': str(row[1]).replace('/', '.'),
                 'days': row[13],
                 'sum': "{:.2f}".format(row[2]),
                 'rate': "{:.2f}".format(row[3]),
                 'interest': "{:.2f}".format(row[12]),
                 'total': "{:.2f}".format(row[5])}
            )
            ptotal += row[3]
            itotal += row[12]
            stotal += row[5]

        guarantor1_name = ""
        guarantor2_name = ""
        guarantor1_location = ""
        guarantor2_location = ""
        guarantor1_phone_num = ""
        guarantor2_phone_num = ""
        if guarantors:
            guarantor1_name = str(
                guarantors[0].last_name) + " " + str(guarantors[0].first_name)
            guarantor1_location = guarantors[0].location
            guarantor1_phone_num = guarantors[0].phone_num
            if len(guarantors) > 1:
                guarantor2_name = str(
                    guarantors[1].last_name) + " " + str(guarantors[1].first_name)
                guarantor2_location = guarantors[1].location
                guarantor2_phone_num = guarantors[1].phone_num

        if client.idnp:
            if client.person == "Physical":
                context = {
                    'id': pk,
                    'contract_date': str(date.fromisoformat(str(contract.date_created)).strftime("%d.%m.%Y")),
                    'first_name': client.first_name,
                    'last_name': client.last_name,
                    'location': client.location,
                    'idnp': client.idnp,
                    'id_card_nr': client.id_card_nr,
                    'id_card_date': str(date.fromisoformat(
                        str(client.id_card_date)).strftime("%d.%m.%Y")
                    ),
                    'id_card_office': client.id_card_office,
                    'birth_date': str(date.fromisoformat(
                        str(client.birth_date)).strftime("%d.%m.%Y")
                    ),
                    'loan_sum': "{:.2f}".format(contract.loan),
                    'scope': contract.scope,
                    'phone_num': client.phone_num,
                    'months': contract.months,
                    'grace_period': contract.grace_period,
                    'date': date.today().strftime("%d.%m.%y"),
                    'guarantor1_name': guarantor1_name,
                    'guarantor2_name': guarantor2_name,
                    'guarantor1_location': guarantor1_location,
                    'guarantor2_location': guarantor2_location,
                    'guarantor1_phone_num': guarantor1_phone_num,
                    'guarantor2_phone_num': guarantor2_phone_num
                }
            else:
                context = {
                    'id': pk,
                    'contract_date': str(date.fromisoformat(str(contract.date_created)).strftime("%d.%m.%Y")),
                    'first_name': client.first_name,
                    'last_name': client.last_name,
                    'location': client.location,
                    'idnp': client.idnp,
                    'id_card_nr': client.id_card_nr,
                    'id_card_date': str(date.fromisoformat(
                        str(client.id_card_date)).strftime("%d.%m.%Y")
                    ),
                    'id_card_office': client.id_card_office,
                    'loan_sum': "{:.2f}".format(contract.loan),
                    'scope': contract.scope,
                    'phone_num': client.phone_num,
                    'months': contract.months,
                    'grace_period': contract.grace_period,
                    'date': date.today().strftime("%d.%m.%y"),
                    'guarantor1_name': guarantor1_name,
                    'guarantor2_name': guarantor2_name,
                    'guarantor1_location': guarantor1_location,
                    'guarantor2_location': guarantor2_location,
                    'guarantor1_phone_num': guarantor1_phone_num,
                    'guarantor2_phone_num': guarantor2_phone_num
                }
            document.render(context)
            buffer = BytesIO()
            document.save(buffer)
            buffer.seek(0)
            response = StreamingHttpResponse(
                streaming_content=buffer,  # use the stream's content
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
            response['Content-Disposition'] = 'attachment;filename=Chestionar_' + \
                str(contract.id) + '.docx'
            response["Content-Encoding"] = 'UTF-8'
            return response
        else:
            messages.error(request, "Client lacks info")
            return redirect('/view_contract/' + str(pk) + '/')


class ExportDocxPaymentsMDL(View):
    def get(self, request, pk, *args, **kwargs):
        # create an empty document object
        directory = os.getcwd()
        global production
        if production:
            document = DocxTemplate(
                directory + '/CRM/accounts/templates/pdf/payments_mdl.docx')
        else:
            document = DocxTemplate(
                directory + '/accounts/templates/pdf/payments_mdl.docx')
        contract = Contract.objects.get(id=pk)
        client = Client.objects.get(id=contract.client_id)
        payments = Payment.objects.filter(contract_id=pk)
        table, profile, penalties = create_spread_sheet(
            str(contract.date_created).split(" ")[0],
            contract.months,
            contract.annual_payments,
            float(contract.loan),
            float(contract.interest_rate),
            contract.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract.month_sum),
            contract.grace_period,
            float(contract.fee_issue or 0),
            int(contract.fee_issue_months or 1),
            float(contract.comission or 0)
        )

        tbl_contents = []
        k = 1
        total_p = 0
        for row in payments:
            tbl_contents.append(
                {
                    'nr': k,
                    'date': row.date_paid.strftime("%d/%m/%y"),
                    'sum': "{:6,.2f}".format((float(round(row.sum / row.exchange_rate, 2))))
                }
            )
            total_p += float(round(row.sum / row.exchange_rate, 2))
            k += 1

        context = {
            'id': pk,
            'date': profile[0][0],
            'first_name': client.first_name,
            'last_name': client.last_name,
            'principal': "{:6,.2f}".format(contract.loan),
            'interest': "{:6,.2f}".format(profile[0][11]),
            'total': "{:6,.2f}".format(profile[0][11] + float(contract.loan)),
            'pay': "{:6,.2f}".format(total_p),
            'balance': "{:6,.2f}".format(total_p - (profile[0][11] + float(contract.loan)) - profile[0][9]),
            'penalty': "{:6,.2f}".format(profile[0][9]),
            'tbl_contents': tbl_contents
        }
        document.render(context)
        buffer = BytesIO()
        document.save(buffer)
        buffer.seek(0)
        response = StreamingHttpResponse(
            streaming_content=buffer,  # use the stream's content
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        response['Content-Disposition'] = 'attachment;filename=Payments_' + \
            str(contract.id) + '.docx'
        response["Content-Encoding"] = 'UTF-8'
        return response


class ExportDocx_Fide(View):
    def get(self, request, pk, *args, **kwargs):
        # create an empty document object
        directory = os.getcwd()
        global production
        if production:
            document = DocxTemplate(
                directory + '/CRM/accounts/templates/pdf/template2.docx')
        else:
            document = DocxTemplate(
                directory + '/accounts/templates/pdf/template2.docx')
        guarantor = Fidejusor.objects.get(id=pk)
        contract = guarantor.contract
        client = contract.client
        print(guarantor.num)

        payments = Payment.objects.filter(contract_id=contract.id)

        table, profile, penalties = create_spread_sheet(
            str(contract.date_created).split(" ")[0],
            contract.months,
            contract.annual_payments,
            float(contract.loan),
            float(contract.interest_rate),
            contract.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract.month_sum),
            contract.grace_period,
            float(contract.fee_issue or 0),
            int(contract.fee_issue_months or 1),
            float(contract.comission or 0)
        )

        context = {
            'id': str(contract.id) + '-' + str(guarantor.num),
            'contract_date': str(date.fromisoformat(str(contract.date_created)).strftime("%d.%m.%Y")),
            'first_name': guarantor.first_name,
            'last_name': guarantor.last_name,
            'location': guarantor.location,
            'id_card_nr': guarantor.id_card_nr,
            'id_card_office': guarantor.id_card_office,
            'id_card_date': str(date.fromisoformat(str(guarantor.id_card_date)).strftime("%d.%m.%Y")),
            'idnp': guarantor.idnp,
            'birth_date': str(date.fromisoformat(str(guarantor.birth_date)).strftime("%d.%m.%Y")),
            'c_id': str(contract.id),
            'c_first_name': client.first_name,
            'c_last_name': client.last_name,
            'c_location': client.location,
            'c_id_card_nr': client.id_card_nr,
            'c_id_card_office': client.id_card_office,
            'c_id_card_date': str(date.fromisoformat(str(client.id_card_date)).strftime("%d.%m.%Y")),
            'c_idnp': client.idnp,
            'c_birth_date': str(date.fromisoformat(str(client.birth_date)).strftime("%d.%m.%Y")),
            'loan_sum': "{:.2f}".format(contract.loan),
            'loan_sum_verbose': verbose_loan(round(float(contract.loan), 2)),
            'last_payment_date': str(table[len(table)-1][1]).replace('/', '.'),
            'phone_num': guarantor.phone_num,
        }
        document.render(context)

        # save document info
        buffer = BytesIO()
        document.save(buffer)  # save your memory stream
        buffer.seek(0)  # rewind the stream

        # put them to streaming content response
        # within docx content_type
        response = StreamingHttpResponse(
            streaming_content=buffer,  # use the stream's content
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

        response['Content-Disposition'] = 'attachment;filename=ContractFidejusiune_' + \
            str(contract.id) + '-' + str(guarantor.num) + '.docx'
        response["Content-Encoding"] = 'UTF-8'

        return response


class ExportDocx_Lien(View):
    def get(self, request, pk, *args, **kwargs):
        # create an empty document object
        directory = os.getcwd()
        global production
        if production:
            document = DocxTemplate(
                directory + '/CRM/accounts/templates/pdf/template_lien.docx')
        else:
            document = DocxTemplate(
                directory + '/accounts/templates/pdf/template_lien.docx')
        lienholder = Gajist.objects.get(id=pk)
        contract = lienholder.contract
        client = contract.client

        context = {
            'id': str(contract.id) + '-' + str(lienholder.num),
            'first_name': lienholder.first_name,
            'last_name': lienholder.last_name,
            'domiciliul': lienholder.domiciliul,
            'idnp': lienholder.idnp,
            'model': lienholder.model,
            'marca': lienholder.marca,
            'nr_de_inmatriculare': lienholder.nr_de_inmatriculare,
            'anul_fabricatiei': lienholder.anul_fabricatiei,
            'nr_caroserie': lienholder.nr_caroserie,
            'culoare': lienholder.culoare,
            'seria_certificat': lienholder.seria_certificat,
            'nr_certificat': lienholder.nr_certificat,
            'data_emiterii_certificat': lienholder.data_emiterii_certificat,
            'valoarea_de_gaj': "{:.2f}".format(lienholder.valoarea_de_gaj),
            'valoarea_de_gaj_verbose': verbose_loan(round(float(lienholder.valoarea_de_gaj), 2)),
            'phone_num': lienholder.phone_num,
            'email': lienholder.email,
            'client_first_name': client.first_name,
            'client_last_name': client.last_name,
            'contract_num': contract.id,
            'contract_date': str(date.fromisoformat(str(contract.date_created)).strftime("%d.%m.%Y")),
        }
        document.render(context)

        # save document info
        buffer = BytesIO()
        document.save(buffer)  # save your memory stream
        buffer.seek(0)  # rewind the stream

        # put them to streaming content response
        # within docx content_type
        response = StreamingHttpResponse(
            streaming_content=buffer,  # use the stream's content
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

        response['Content-Disposition'] = 'attachment;filename=ContractGaj_' + \
            str(contract.id) + '-' + str(lienholder.num) + '.docx'
        response["Content-Encoding"] = 'UTF-8'

        return response


class ExportDocx_LienAviz(View):
    def get(self, request, pk, *args, **kwargs):
        # create an empty document object
        directory = os.getcwd()
        global production
        if production:
            document = DocxTemplate(
                directory + '/CRM/accounts/templates/pdf/template_lien_aviz.docx')
        else:
            document = DocxTemplate(
                directory + '/accounts/templates/pdf/template_lien_aviz.docx')
        lienholder = Gajist.objects.get(id=pk)
        contract = lienholder.contract
        client = contract.client

        payments = Payment.objects.filter(contract_id=contract.id)

        table, profile, penalties = create_spread_sheet(
            str(contract.date_created).split(" ")[0],
            contract.months,
            contract.annual_payments,
            float(contract.loan),
            float(contract.interest_rate),
            contract.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract.month_sum),
            contract.grace_period,
            float(contract.fee_issue or 0),
            int(contract.fee_issue_months or 1),
            float(contract.comission or 0)
        )

        context = {
            'id': str(contract.id) + '-' + str(lienholder.num),
            'first_name': lienholder.first_name,
            'last_name': lienholder.last_name,
            'domiciliul': lienholder.domiciliul,
            'idnp': lienholder.idnp,
            'model': lienholder.model,
            'marca': lienholder.marca,
            'nr_de_inmatriculare': lienholder.nr_de_inmatriculare,
            'anul_fabricatiei': lienholder.anul_fabricatiei,
            'nr_caroserie': lienholder.nr_caroserie,
            'culoare': lienholder.culoare,
            'seria_certificat': lienholder.seria_certificat,
            'nr_certificat': lienholder.nr_certificat,
            'data_emiterii_certificat': lienholder.data_emiterii_certificat,
            'valoarea_de_gaj': "{:.2f}".format(lienholder.valoarea_de_gaj),
            'valoarea_de_gaj_verbose': verbose_loan(round(float(lienholder.valoarea_de_gaj), 2)),
            'phone_num': lienholder.phone_num,
            'email': lienholder.email,
            'client_first_name': client.first_name,
            'client_last_name': client.last_name,
            'contract_num': contract.id,
            'contract_date': str(date.fromisoformat(str(contract.date_created)).strftime("%d.%m.%Y")),
            'loan_sum': "{:.2f}".format(contract.loan),
            'contract_months': contract.months,
            'last_payment_date': str(table[len(table)-1][1]).replace('/', '.'),
        }
        document.render(context)

        # save document info
        buffer = BytesIO()
        document.save(buffer)  # save your memory stream
        buffer.seek(0)  # rewind the stream

        # put them to streaming content response
        # within docx content_type
        response = StreamingHttpResponse(
            streaming_content=buffer,  # use the stream's content
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

        response['Content-Disposition'] = 'attachment;filename=AvizGaj_' + \
            str(contract.id) + '-' + str(lienholder.num) + '.docx'
        response["Content-Encoding"] = 'UTF-8'

        return response


class GeneratePDF(View):
    def get(self, request, pk, *args, **kwargs):
        directory = os.getcwd()
        global production
        if production:
            document = DocxTemplate(
                directory + '/CRM/accounts/templates/pdf/payments.pdf')
        else:
            document = DocxTemplate(
                directory + '/accounts/templates/pdf/payments.pdf')
        contract = Contract.objects.get(id=pk)
        client = Client.objects.get(id=contract.client_id)
        payments = Payment.objects.filter(contract_id=pk)
        count = 0

        guarantors = Fidejusor.objects.filter(contract=contract)

        all_payments = []
        for payment in payments:
            count += 1
            if contract.currency == "EUR":
                all_payments.append([
                    payment,
                    count,
                    True,
                    round(float(payment.exchange_rate), 2),
                    round(float(payment.sum / payment.exchange_rate), 2),
                ])
            else:
                all_payments.append([
                    payment,
                    count,
                    False
                ])

        total_mdl = 0
        total_eur = 0
        for payment in payments:
            total_mdl += payment.sum
            total_eur += payment.sum / payment.exchange_rate
        totals = (round(float(total_mdl), 2),
                  round(float(total_eur), 2))
        table, profile, penalties = create_spread_sheet(
            str(contract.date_created).split(" ")[0],
            contract.months,
            contract.annual_payments,
            float(contract.loan),
            float(contract.interest_rate),
            contract.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract.month_sum),
            contract.grace_period,
            float(contract.fee_issue or 0),
            int(contract.fee_issue_months or 1),
            float(contract.comission or 0)
        )
        context = {'client': client,
                   'table': table,
                   'profile': profile,
                   'penalties': penalties,
                   'all_payments': all_payments,
                   'contract': contract,
                   'total_payments': totals,
                   'guarantors': guarantors
                   }
        html = template.render(context)
        pdf = render_to_pdf('pdf/payments.html', context)
        # pdf = render_to_pdf(directory + '/CRM/accounts/templates/pdf/payments.pdf', context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Payments_" + str(contract.id) + ".pdf"
            content = "inline; filename='%s'" % (filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" % (filename)
            response['Content-Disposition'] = content
            return response
        return HttpResponse("Not found")


@ unauthenticated_user
def registerPage(request):
    form = CreateUserForm()
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')

            messages.success(request, 'Account was created for ' + username)

            return redirect('login')

    context = {'form': form}
    return render(request, 'accounts/register.html', context)


@ unauthenticated_user
def loginPage(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.info(request, 'Username OR password is incorrect')

    context = {}
    return render(request, 'accounts/login.html', context)


def logoutUser(request):
    logout(request)
    return redirect('login')


@ login_required(login_url='login')
@ admin_only
def home(request):
    clients = Client.objects.all()
    total_clients = clients.count()
    myFilter = ClientFilter(request.GET, queryset=clients)
    clients = myFilter.qs

    context = {'clients': clients,
               'total_clients': total_clients,
               'myFilter': myFilter}

    return render(request, 'accounts/my_dashboard.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def arrears(request):
    contracts_list = Contract.objects.all()
    proc_date = date.today()
    myFilter = ArrearsFilter(request.GET, queryset=contracts_list)
    person = 'All'
    if request.GET & ArrearsForm.base_fields.keys():
        form = ArrearsForm(request.GET)
        proc_date = date.fromisoformat(form.data.get('date'))
        person = form.data.get('person')
    else:
        form = ArrearsForm(initial={
            'date': proc_date
        })
    if myFilter.is_valid():
        status = form.data.get('status')
    contracts_list = myFilter.qs
    contracts = []
    profile = [0, 0, 0, 0, 0]
    filter = [1, 0, 0, 0, 0]
    closed = False
    for contract_data in contracts_list:
        payments = Payment.objects.filter(contract_id=contract_data.id)
        client = Client.objects.get(id=contract_data.client_id)
        if status == 'Closed':
            closed = True
        if status != 'Closed' and contract_data.status == 'Closed':
            continue
        if person != 'All' and person != client.person:
            continue
        if contract_data.date_created > proc_date:
            continue
        contract_info = create_spread_sheet(
            str(contract_data.date_created).split(" ")[0],
            contract_data.months,
            contract_data.annual_payments,
            float(contract_data.loan),
            float(contract_data.interest_rate),
            contract_data.calc_method,
            str(proc_date),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract_data.month_sum),
            contract_data.grace_period,
            float(contract_data.fee_issue or 0),
            int(contract_data.fee_issue_months or 1),
            float(contract_data.comission or 0)
        )
        if client.person == "Physical":
            name = str(client.last_name) + " " + str(client.first_name)
        else:
            name = str(client.company_name)
        contracts.append([
            contract_data.id,
            name,
            contract_info[1],
            contract_data.client_id,
            contract_data.date_created,
            contract_data.date_closed,
            contract_info[1][0][29]
        ])
    count = contracts.__len__()
    for contract in contracts:
        profile[0] += contract[2][0][10]
        profile[1] += contract[2][0][12]
        if contract[2][0][21] > 30:
            profile[2] += contract[2][0][4]
        profile[4] += contract[2][0][5]
    if profile[1] > 0:
        profile[3] = round((profile[2] / profile[1]) * 100, 2)
    for i in range(len(profile)):
        profile[i] = round(profile[i], 2)

    context = {'contracts': contracts,
               'profile': profile,
               'filter': filter,
               'myFilter': myFilter,
               'count': count,
               'additionalForm': form,
               'closed': closed,
               }
    return render(request, 'accounts/arrears.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def payments(request):
    end_date = date.today()
    start_date = date.fromisocalendar(2001, 1, 1)
    if request.GET:
        form = PaymentsForm(request.GET)
        end_date = date.fromisoformat(form.data.get('end_date'))
        start_date = date.fromisoformat(form.data.get('start_date'))
    else:
        form = PaymentsForm(initial={
            'start_date': date.fromisocalendar(2001, 1, 1),
            'end_date': end_date
        })

    payments = Payment.objects.filter(
        date_paid__range=[start_date, end_date])
    print(start_date)
    print(end_date)
    print(payments)

    context = {
        'payments': payments,
        'form': form
    }
    return render(request, 'accounts/payments.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def client(request, pk_test):
    client = Client.objects.get(id=pk_test)
    contracts_list = client.contract_set.all()
    contracts_count = contracts_list.count()
    contracts = []
    for contract_data in contracts_list:
        payments = Payment.objects.filter(contract_id=contract_data.id)
        contracts.append([
            contract_data.id,
            contract_data.date_created,
            contract_data.product,
            contract_data.loan,
            contract_data.status,
            create_spread_sheet(
                str(contract_data.date_created).split(" ")[0],
                contract_data.months,
                contract_data.annual_payments,
                float(contract_data.loan),
                float(contract_data.interest_rate),
                contract_data.calc_method,
                str(date.today()),
                [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
                float(contract_data.month_sum),
                contract_data.grace_period,
                float(contract_data.fee_issue or 0),
                int(contract_data.fee_issue_months or 1),
                float(contract_data.comission or 0)
            )[1]
        ])
    context = {'client': client, 'contracts': contracts,
               'contracts_count': contracts_count}
    return render(request, 'accounts/client.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def createContract(request, pk):
    client = Client.objects.get(id=pk)
    ContractForms = modelform_factory(Contract, form=ContractForm)
    form = ContractForms(initial={'client': client})
    client_id = client.pk
    created = False
    if request.method == 'POST':
        form = ContractForm(request.POST)
        if form.is_valid():
            form.save()
            contract = Contract.objects.get(id=form.data.get('id'))
            return redirect('/view_contract/' + str(contract.id) + '/')
    context = {'form': form, 'created': created}
    return render(request, 'accounts/contract_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def updateContract(request, pk):
    contract = Contract.objects.get(id=pk)
    form = ContractForm(instance=contract)
    created = True
    guarantors = contract.fidejusor_set.all()
    lienholders = contract.gajist_set.all()
    count = guarantors.count()
    if request.method == 'POST':
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            form.save()
            return redirect(f'/client/{contract.client_id}/')
    context = {'form': form,
               'created': created,
               'guarantors': guarantors,
               'lienholders': lienholders,
               'count': count,
               'contract': contract}
    return render(request, 'accounts/contract_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def viewContract(request, pk):
    contract = Contract.objects.get(id=pk)
    client = Client.objects.get(id=contract.client_id)
    payments = Payment.objects.filter(contract_id=pk)
    count = 0
    closed = False
    if contract.status == 'Closed':
        closed = True
    guarantors = Fidejusor.objects.filter(contract=contract)
    lienholders = Gajist.objects.filter(contract=contract)

    # Per-payment distribution breakdown (Penalitate -> Dobandă -> Corp), mirroring
    # the priority cascade used both by create_spread_sheet and the index-eur calculator.
    sorted_payments = list(payments.order_by('date_paid', 'id'))
    payment_breakdown = {}
    prev_principal_paid = 0.0
    prev_interest_paid = 0.0
    prev_penalty_paid = 0.0
    prev_commission_paid = 0.0
    for i, sorted_payment in enumerate(sorted_payments):
        subset = sorted_payments[:i + 1]
        _, prefix_profile, _ = create_spread_sheet(
            str(contract.date_created).split(" ")[0],
            contract.months,
            contract.annual_payments,
            float(contract.loan),
            float(contract.interest_rate),
            contract.calc_method,
            str(sorted_payment.date_paid),
            [(float(round(pp.sum / pp.exchange_rate, 2)), pp.date_paid, float(pp.penalty or 0)) for pp in subset],
            float(contract.month_sum),
            contract.grace_period,
            float(contract.fee_issue or 0),
            int(contract.fee_issue_months or 1),
            float(contract.comission or 0)
        )
        row = prefix_profile[0]
        payment_breakdown[sorted_payment.id] = {
            'to_penalty': round(row[9] - prev_penalty_paid, 2),
            'to_commission': round(row[17] - prev_commission_paid, 2),
            'to_interest': round(row[8] - prev_interest_paid, 2),
            'to_principal': round(row[7] - prev_principal_paid, 2),
            'corp_remaining': row[12],
            'interest_remaining': row[13],
        }
        prev_penalty_paid = row[9]
        prev_commission_paid = row[17]
        prev_interest_paid = row[8]
        prev_principal_paid = row[7]

    all_payments = []
    for payment in payments:
        count += 1
        breakdown = payment_breakdown.get(payment.id, {})
        if contract.currency == "EUR":
            all_payments.append([
                payment,
                count,
                True,
                round(float(payment.exchange_rate), 2),
                round(float(payment.sum / payment.exchange_rate), 2),
                breakdown,
            ])
        else:
            all_payments.append([
                payment,
                count,
                False,
                None,
                None,
                breakdown,
            ])

    total_mdl = 0
    total_eur = 0
    for payment in payments:
        total_mdl += payment.sum
        total_eur += payment.sum / payment.exchange_rate
    totals = (round(float(total_mdl), 2),
              round(float(total_eur), 2))

    table, profile, penalties = create_spread_sheet(
        str(contract.date_created).split(" ")[0],
        contract.months,
        contract.annual_payments,
        float(contract.loan),
        float(contract.interest_rate),
        contract.calc_method,
        str(date.today()),
        [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0))
         for payment in payments],
        float(contract.month_sum),
        contract.grace_period,
        float(contract.fee_issue or 0),
        int(contract.fee_issue_months or 1),
        float(contract.comission or 0)
    )
    comments = contract.comments
    commentForm = CommentsForm(instance=contract, initial={
        'comments': comments
    })

    if request.method == 'POST':
        commentForm = CommentsForm(request.POST, instance=contract)
        form = ContractForm(request.POST, instance=contract)
        if commentForm.is_valid():
            commentForm.save()
        if form.is_valid():
            form.save()
            return redirect('/')

    PaymentForms = modelform_factory(Payment, form=PaymentForm)
    payment_form = PaymentForms(initial={'contract': contract, 'date_paid': date.today()})

    tracker_rows = []
    for _row in table:
        _is_ca     = _row[0] == 'CA'
        _plan_corp = _row[3]
        _plan_dob  = _row[4]
        _plan_com  = _row[6]
        _paid_corp = _row[7]
        _paid_dob  = _row[8]
        _paid_com  = _row[9]
        _total_paid = round(_paid_corp + _paid_dob + _paid_com, 2)
        if _is_ca:
            _corp_done = True
            _dob_done  = True
            _fully_closed = _paid_com >= _plan_com - 0.005
            _partial      = _paid_com > 0.005 and not _fully_closed
        else:
            _corp_done    = _paid_corp >= _plan_corp - 0.005
            _dob_done     = _paid_dob  >= _plan_dob  - 0.005
            _fully_closed = _corp_done and _dob_done
            _partial      = _total_paid > 0.005 and not _fully_closed
        if _fully_closed:
            _status = 'closed'
        elif _partial:
            _status = 'partial'
        else:
            _status = 'pending'
        tracker_rows.append({
            'n':          _row[0],
            'date':       _row[1],
            'is_ca':      _is_ca,
            'plan_corp':  _plan_corp,
            'paid_corp':  _paid_corp,
            'corp_done':  _corp_done,
            'plan_dob':   _plan_dob,
            'paid_dob':   _paid_dob,
            'dob_done':   _dob_done,
            'plan_com':   _plan_com,
            'paid_com':   _paid_com,
            'total_paid': _total_paid,
            'status':     _status,
        })
    tracker_total = len(tracker_rows)
    tracker_paid  = sum(1 for r in tracker_rows if r['status'] == 'closed')
    tracker_total_paid = round(sum(r['total_paid'] for r in tracker_rows), 2)

    context = {'client': client,
               'table': table,
               'profile': profile,
               'penalties': penalties,
               'all_payments': all_payments,
               'contract': contract,
               'total_payments': totals,
               'guarantors': guarantors,
               'closed': closed,
               'commentForm': commentForm,
               'lienholders': lienholders,
               'payment_form': payment_form,
               'tracker_rows': tracker_rows,
               'tracker_paid': tracker_paid,
               'tracker_total': tracker_total,
               'tracker_total_paid': tracker_total_paid,
               }
    return render(request, 'accounts/contract_table.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deleteContract(request, pk):
    contract = Contract.objects.get(id=pk)
    client = contract.client
    client_id = client.pk
    if request.method == "POST":
        contract.delete()
        return redirect(f'/client/{client.id}/')
    context = {'item': contract}
    return render(request, 'accounts/delete_contract.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def createClient(request):
    PhysicalClientForms = modelform_factory(
        Client, form=PhysicalClientForm, fields='__all__')
    physical_form = PhysicalClientForms(initial={'person': 'Physical'})
    JuridicalClientForms = modelform_factory(
        Client, form=JuridicalClientForm, fields='__all__')
    juridical_form = JuridicalClientForms(initial={'person': 'Juridical'})
    if request.method == 'POST':
        physical_form = PhysicalClientForm(request.POST)
        juridical_form = JuridicalClientForm(request.POST)
        if physical_form.is_valid():
            physical_form.save()
            client = Client.objects.get(idnp=physical_form.data.get('idnp'))
            return redirect('/client/' + str(client.id) + '/')
        if juridical_form.is_valid():
            juridical_form.save()
            client = Client.objects.get(idnp=juridical_form.data.get('idnp'))
            return redirect('/client/' + str(client.id) + '/')
    context = {
        'physical_form': physical_form,
        'juridical_form': juridical_form
    }
    return render(request, 'accounts/new_client.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def updateClient(request, pk):
    client = Client.objects.get(id=pk)
    if client.person == "Physical":
        form = PhysicalClientForm(instance=client)
    else:
        form = JuridicalClientForm(instance=client)
    if request.method == 'POST':
        physical_form = PhysicalClientForm(request.POST, instance=client)
        juridical_form = JuridicalClientForm(request.POST, instance=client)
        if physical_form.is_valid():
            physical_form.save()
            print('phys saved')
            return redirect(f'/client/{client.id}/')
        elif juridical_form.is_valid():
            juridical_form.save()
            print('jur saved')
            return redirect(f'/client/{client.id}/')

    context = {'form': form}
    return render(request, 'accounts/new_client.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deleteClient(request, pk):
    client = Client.objects.get(id=pk)
    if request.method == "POST":
        client.delete()
        return redirect('/')

    context = {'item': client}
    return render(request, 'accounts/delete_client.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def createPayment(request, pk):
    contract = Contract.objects.get(id=pk)
    client = contract.client
    PaymentForms = modelform_factory(Payment, form=PaymentForm)
    form = PaymentForms(initial={'contract': contract})
    contract_id = contract.pk
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        new_form = Payment()
        if form.is_valid():
            form.save()
            return redirect('/view_contract/' + str(contract_id) + '/')
    context = {'form': form, 'contract': contract, 'client': client}
    return render(request, 'accounts/payment_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def updatePayment(request, pk):
    payment = Payment.objects.get(id=pk)
    form = PaymentForm(instance=payment)
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            return redirect(f'/view_contract/{payment.contract_id}/')
    context = {'form': form}
    return render(request, 'accounts/payment_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deletePayment(request, pk):
    payment = Payment.objects.get(id=pk)
    if request.method == "POST":
        payment.delete()
        return redirect(f'/view_contract/{payment.contract_id}/')

    context = {'item': payment}
    return render(request, 'accounts/delete_payment.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def createPenaltyWaive(request, pk, pj):
    waive = PenaltyWaive(contract_id=pk, month=pj)
    waive.save()
    return redirect('/view_contract/' + str(waive.contract_id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def createInterestWaive(request, pk, pj):
    waive = InterestWaive(contract_id=pk, month=pj)
    waive.save()
    return redirect('/view_contract/' + str(waive.contract_id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deletePenaltyWaive(request, pk, pj):
    waive = PenaltyWaive.objects.filter(contract_id=pk, month=pj)
    id = pk
    waive.delete()
    return redirect('/view_contract/' + str(id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deleteInterestWaive(request, pk, pj):
    waive = InterestWaive.objects.filter(contract_id=pk, month=pj)
    id = pk
    waive.delete()
    return redirect('/view_contract/' + str(id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def waiveAllInterest(request, pk):
    contract = Contract.objects.get(id=pk)
    for i in range(contract.months):
        waive = InterestWaive(contract_id=pk, month=i+1)
        waive.save()
    return redirect('/view_contract/' + str(waive.contract_id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def waiveAllPenalty(request, pk):
    contract = Contract.objects.get(id=pk)
    for i in range(contract.months):
        waive = PenaltyWaive(contract_id=pk, month=i+1)
        waive.save()
    return redirect('/view_contract/' + str(waive.contract_id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deleteAllPenalty(request, pk):
    contract = Contract.objects.get(id=pk)
    for i in range(contract.months):
        waive = PenaltyWaive.objects.filter(contract_id=pk, month=i+1)
        id = pk
        waive.delete()
    return redirect('/view_contract/' + str(id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deleteAllInterest(request, pk):
    contract = Contract.objects.get(id=pk)
    for i in range(contract.months):
        waive = InterestWaive.objects.filter(contract_id=pk, month=i+1)
        id = pk
        waive.delete()
    return redirect('/view_contract/' + str(id) + '/')


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def createGuarantor(request, pk):
    contract = Contract.objects.get(id=pk)
    GuarantorForms = modelform_factory(Fidejusor, form=GuarantorForm)
    form = GuarantorForms(initial={'contract': contract})
    contract_id = contract.pk
    if request.method == 'POST':
        form = GuarantorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/update_contract/' + str(contract_id) + '/')
    context = {'form': form}
    return render(request, 'accounts/guarantor_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def createLienholder(request, pk):
    contract = Contract.objects.get(id=pk)
    GajistForms = modelform_factory(Gajist, form=GajistForm)
    form = GajistForms(initial={'contract': contract})
    contract_id = contract.pk
    if request.method == 'POST':
        form = GajistForms(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/update_contract/' + str(contract_id) + '/')
    context = {'form': form}
    return render(request, 'accounts/lienholder_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deleteGuarantor(request, pk):
    guarantor = Fidejusor.objects.get(id=pk)
    if request.method == "POST":
        guarantor.delete()
        return redirect(f'/update_contract/{guarantor.contract.id}/')

    context = {'item': guarantor}
    return render(request, 'accounts/delete_guarantor.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def deleteLienholder(request, pk):
    lienholder = Gajist.objects.get(id=pk)
    if request.method == "POST":
        lienholder.delete()
        return redirect(f'/update_contract/{lienholder.contract.id}/')
    context = {'item': lienholder}
    return render(request, 'accounts/delete_lienholder.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def updateGuarantor(request, pk):
    guarantor = Fidejusor.objects.get(id=pk)
    form = GuarantorForm(instance=guarantor)
    if request.method == 'POST':
        form = GuarantorForm(request.POST, instance=guarantor)
        if form.is_valid():
            form.save()
            return redirect(f'/update_contract/{guarantor.contract.id}/')
    context = {'form': form}
    return render(request, 'accounts/guarantor_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def updateLienholder(request, pk):
    lienholder = Gajist.objects.get(id=pk)
    form = GajistForm(instance=lienholder)
    if request.method == 'POST':
        form = GajistForm(request.POST, instance=lienholder)
        if form.is_valid():
            form.save()
            return redirect(f'/update_contract/{lienholder.contract.id}/')
    context = {'form': form}
    return render(request, 'accounts/lienholder_form.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def prognosis(request, pk, pj, pl):
    if pl == 'All':
        contracts_list = Contract.objects.filter(status='Active')
    else:
        contracts_list = Contract.objects.filter(status='Active', product=pl)
    start_date = datetime.fromisoformat(pk).date()
    end_date = datetime.fromisoformat(pj).date()
    form = PrognosisForm(initial={'start_date': start_date,
                                  'end_date': end_date,
                                  'product': pl
                                  })
    full = []
    total = [0, 0, 0, 0]
    for contract_data in contracts_list:
        payments = Payment.objects.filter(contract_id=contract_data.id)
        client = Client.objects.get(id=contract_data.client_id)
        table, profile, penalties = create_spread_sheet(
            str(contract_data.date_created).split(" ")[0],
            contract_data.months,
            contract_data.annual_payments,
            float(contract_data.loan),
            float(contract_data.interest_rate),
            contract_data.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract_data.month_sum),
            contract_data.grace_period,
            float(contract_data.fee_issue or 0),
            int(contract_data.fee_issue_months or 1),
            float(contract_data.comission or 0)
        )
        state = False
        data = [0, 0, 0]
        for tab in table:
            if tab[14] > start_date and tab[14] < end_date:
                state = True
                data[0] += tab[3]
                total[0] += tab[3]
                data[1] += tab[4]
                total[1] += tab[4]
                data[2] += tab[5]
                total[2] += tab[5]
        for i in range(len(data)):
            data[i] = round(data[i], 2)
        if state:
            full.append([
                str(contract_data.id),
                str(client.last_name) + ' ' + str(client.first_name),
                data
            ])
            total[3] += 1
    for i in range(len(total)):
        total[i] = round(total[i], 2)
    if request.method == 'POST':
        form = PrognosisForm(request.POST)
        if form.is_valid():
            start = form.data.get('start_date')
            end = form.data.get('end_date')
            product = form.data.get('product')
            return redirect('/prognosis/' + start + '/' + end + '/' + product + '/')
    context = {
        'data': full,
        'total': total,
        'form': form
    }
    return render(request, 'accounts/prognosis.html', context)


@ login_required(login_url='login')
@ allowed_users(allowed_roles=['admin'])
def report(request, pk, pj):
    if pj == 'All':
        contracts_list = Contract.objects.all()
    else:
        contracts_list = Contract.objects.filter(product=pl)
    start_date = datetime.fromisoformat(pk).date()
    form = ReportForm(initial={'start_date': start_date,
                               'product': pj
                               })
    total = [0, 0, 0]
    data = {
        'contract_nr': 0,
        'disbursed_amount': 0,
        'interest_paid': 0
    }
    interest_paid = []
    graph_points1 = [0 for _ in range(12)]
    graph_points2 = [0 for _ in range(12)]
    graph_points3 = [0 for _ in range(12)]

    for contract_data in contracts_list:
        payments = Payment.objects.filter(contract_id=contract_data.id)
        client = Client.objects.get(id=contract_data.client_id)
        table, profile, penalties = create_spread_sheet(
            str(contract_data.date_created).split(" ")[0],
            contract_data.months,
            contract_data.annual_payments,
            float(contract_data.loan),
            float(contract_data.interest_rate),
            contract_data.calc_method,
            str(date.today()),
            [(float(round(payment.sum / payment.exchange_rate, 2)), payment.date_paid, float(payment.penalty or 0)) for payment in payments],
            float(contract_data.month_sum),
            contract_data.grace_period,
            float(contract_data.fee_issue or 0),
            int(contract_data.fee_issue_months or 1),
            float(contract_data.comission or 0)
        )

        for month_add in range(12):
            for tab in table:
                if tab[14].month == (start_date + relativedelta(months=month_add)).month and tab[14].year == (start_date + relativedelta(months=month_add)).year:
                    graph_points3[month_add] += float(tab[8])
                    data['interest_paid'] += float(tab[8])
            if contract_data.date_created > start_date and contract_data.date_created < start_date + relativedelta(years=1):
                if contract_data.date_created.month == (start_date + relativedelta(months=month_add)).month:
                    graph_points1[month_add] += 1
                    graph_points2[month_add] += float(contract_data.loan)
                    data['contract_nr'] += 1
                    data['disbursed_amount'] += float(contract_data.loan)

    graph_labels = [
        str((start_date + relativedelta(months=month_nr)).strftime("%b %y")) for month_nr in range(12)
    ]

    graph_data = {
        'labels': graph_labels,
        'points1': graph_points1,
        'points2': graph_points2,
        'points3': graph_points3
    }

    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            start = form.data.get('start_date')
            product = form.data.get('product')
            return redirect('/report/' + start + '/' + product + '/')
    context = {
        'data': data,
        'form': form,
        'graph_data': graph_data
    }
    return render(request, 'accounts/report.html', context)


@login_required(login_url='login')
def bnm_rate(request):
    try:
        today = date.today().strftime('%d.%m.%Y')
        url = f'https://www.bnm.md/ro/export-official-exchange-rates?date={today}'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            text = resp.read().decode('utf-8')
        # Response is semicolon-separated CSV: Valuta;Cod;Abr;Rata;Cursul
        # EUR row example: Euro;978;EUR;1;20,1491
        for line in text.splitlines():
            parts = line.split(';')
            if len(parts) >= 5 and parts[2].strip().strip('"') == 'EUR':
                rate = float(parts[4].strip().replace(',', '.'))
                return JsonResponse({'success': True, 'rate': rate})
        return JsonResponse({'success': False, 'error': 'EUR not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
